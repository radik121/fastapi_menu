from datetime import datetime
from typing import Any, Dict

import openpyxl

from .celery_app import celery_app


@celery_app.task
def save_data_to_xlsx(data: dict):
    file_name = f"{datetime.now().strftime('%d-%m-%Y-%H-%M')}_menu.xlsx"
    file_path = f'task_files/{file_name}'
    save_to_xlsx(data, file_path)
    return {"file_name": file_name}


def get_result(task_id: str):
    task = celery_app.AsyncResult(task_id)
    return task


def save_to_xlsx(data: Dict[Any, Any], file_name: str) -> None:
    book = openpyxl.Workbook()
    sheet = book.active
    menu_row, menu_column, menu_counter = 1, 1, 1
    for menu in data:
        sheet.cell(menu_row, menu_column).value = menu_counter
        sheet.cell(menu_row, menu_column + 1).value = menu["title"]
        sheet.cell(menu_row, menu_column + 2).value = menu["description"]
        menu_counter += 1
        submenu_row = menu_row + 1
        submenu_column = menu_column + 1
        submenu_counter = 1
        if menu["submenus"]:
            for submenu in menu["submenus"]:
                sheet.cell(submenu_row, submenu_column).value = submenu_counter
                sheet.cell(submenu_row, submenu_column + 1).value = submenu["title"]
                sheet.cell(submenu_row, submenu_column + 2).value = submenu["description"]
                submenu_counter += 1
                dish_row = submenu_row + 1
                dish_column = submenu_column + 1
                dish_counter = 1
                if submenu["dishes"]:
                    for dish in submenu["dishes"]:
                        sheet.cell(dish_row, dish_column).value = dish_counter
                        sheet.cell(dish_row, dish_column + 1).value = dish["title"]
                        sheet.cell(dish_row, dish_column + 2).value = dish["description"]
                        sheet.cell(dish_row, dish_column + 3).value = dish["price"]
                        dish_counter += 1
                        dish_row += 1
                submenu_row = dish_row
        menu_row = submenu_row
    book.save(file_name)
    book.close()
